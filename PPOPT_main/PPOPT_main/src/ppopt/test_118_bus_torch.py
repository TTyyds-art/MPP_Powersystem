# %% 0. 导入所需的包
import datetime

import pandapower as pp
import warnings
import time, torch
import numpy as np
from numpy import flatnonzero as find, r_
from torch.utils.data import Dataset, DataLoader

from openpyxl import Workbook
import os
from PPOPT_main.PPOPT_main.src.ppopt.solution import Solution

import sys, os
path_current = '/home/huzuntao/PycharmProjects/MPP_Powersystem/'
path_ = os.getcwd()
if path_current not in sys.path:
    sys.path.insert(1, '/home/huzuntao/PycharmProjects/MPP_Powersystem/')
elif path_ not in sys.path:
    sys.path.insert(1, path_)

from PPOPT_main.PPOPT_main.src.ppopt.mpQCQP_program_0731 import MPQCQP_Program
from PPOPT_main.PPOPT_main.src.ppopt.utils.mpqp_utils_torch import gen_cr_from_active_set_torch, \
    optimal_control_law_torch

# 设定输出警告信息的方式
warnings.filterwarnings("ignore")  # 忽略掉所有警告
# 检查是否存在GPU，如果存在则使用，否则使用CPU
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

class QCQPDataset(Dataset):
    def __init__(self, sorted_data_list):
        # sorted_data_list 是一个排序后的数据列表
        self.data_list = sorted_data_list

    def __len__(self):
        return len(self.data_list)

    def __getitem__(self, idx):
        return self.data_list[idx]

def save_dataset_exl(Add_tot, dataset, idx_act_dict, unique=True, save=False):

    if not save:
        return None
    if unique:
        flag = "U"
    else:
        flag = "C"
    # 获取当前执行的.py文件的名称（不包括扩展名）
    file_name_without_extension = os.path.splitext(os.path.basename(__file__))[0]
    # 根据文件名创建文件夹名称
    folder_name = f"dataset_{file_name_without_extension}"
    # 确保目标目录存在，如果不存在则创建
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
    # 获取当前的日期和时间，格式化它
    current_time = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"dataset_{current_time}_{flag}{Add_tot}.pth"
    full_path = os.path.join(folder_name, filename)
    torch.save(dataset, full_path)  # 保存数据集
    save_dict_to_excel(idx_act_dict, filename=f"idx_act_dict_{current_time}_{flag}{Add_tot}.xlsx", path=folder_name)
    return full_path

def save_dict_to_excel(data_dict, filename="data.xlsx", path="."):
    """
    保存字典到Excel文件。

    参数:
        data_dict (dict): 要保存的字典。
        filename (str, optional): 保存的Excel文件名。默认为"data.xlsx"。
        path (str, optional): 保存的路径。默认为当前目录。
    """
    wb = Workbook()  # 创建一个工作簿
    ws = wb.active  # 获取当前活动工作表

    # 写入标题（字典的键）
    for row_num, key in enumerate(data_dict.keys(), 1):
        if isinstance(key, tuple):
            key_str = ', '.join(map(str, key))
        else:
            key_str = str(key)
        ws.cell(row=row_num, column=1, value=key_str)

    # 写入数据（字典的值）
    for row_num, value in enumerate(data_dict.values(), 1):
        ws.cell(row=row_num, column=2, value=value)

    save_path = os.path.join(path, filename)
    wb.save(save_path)


def find_ineq_del_idx(net, pq_load):
    idx = 0
    idx_ineq_l = {}
    idx_ineq_u = {}
    while idx <= 1:
        if idx == 0:
            mul_factor = 0.8
        else:
            mul_factor = 1.2
        idx += 1
        net.load.p_mw = pq_load[:net.load.shape[0]] * mul_factor
        net.load.q_mvar = pq_load[net.load.shape[0]:] * mul_factor
        try:
            pp.runopp(net, delta=1e-16, RETURN_RAW_DER=1)
        except:
            print(f"Error happens in running the OPF.")
            return None
        new_QCQP = MPQCQP_Program(net=net, lb_loads=pq_load * 0.8,
                                  ub_loads=pq_load * 1.2)
        mu_ = new_QCQP.raw['lmbda']['mu']
        # idx_act = find(mu_ > 1e-3).tolist()
        if mul_factor == 0.8:
            idx_ineq_l = set(find(mu_ <= 1e-3))
        else:
            idx_ineq_u = set(find(mu_ <= 1e-3))
    ineq_idx_del = idx_ineq_l & idx_ineq_u
    num_ineq = new_QCQP.num_inequality_constraints()
    ineq_keep_del = [i for i in range(num_ineq) if i not in ineq_idx_del]

    return ineq_keep_del


def main():
    # %% 1. 118-bus system 跑通

    from pandapower.networks import case118
    net = case118()

    # %% 1.1 使用Pandapower求解该问题
    pp.runopp(net, verbose=False)

    print(f"load: {net.load.p_mw[:5]}")
    temp_1 = np.squeeze(net.res_bus)  # 获得net通过Pandapower计算后的结果
    print(f"bus: \n{temp_1[:5]}")

    # %% 2. 生成当前的critical region
    # 设置开始的时间
    start_time = time.time()
    # 设定随机数种子 numpy
    np.random.seed(0)
    torch.manual_seed(0)
    # 生成新的负载
    p_load_init = net.load.p_mw.copy()
    q_load_init = net.load.q_mvar.copy()

     # %%  3. 生成最终的solution
    # opf_anly_sol = Solution(program=new_QCQP, critical_regions=[cri_region])
    Add_flag = 0
    Add_tot = 400
    error_count = 0
    all_data_lists_dict = {}
    idx_act_dict = {}
    idx_act_tot = {}
    ineq_keep_del = []
    while Add_flag < Add_tot:
        Add_flag += 1
        print(f"Add_flag = {Add_flag}")
        # 生成新的load
        def generate_combined_load(load_init):
            return np.array([val * (0.8 + 0.4 * np.random.rand()) for val in load_init]).reshape([-1, 1])

        combined_init = np.vstack((p_load_init, q_load_init))
        pq_load = generate_combined_load(combined_init)
        if Add_flag == 1:
            ineq_keep_del = find_ineq_del_idx(net, pq_load=r_[p_load_init, q_load_init])
            ineq_keep_del.append(590)
            print(f"ineq_keep_del = {ineq_keep_del}")

        # print(f'pq_load  = {pq_load.squeeze(axis=1)[:5]}')
        # 对于新的pq_load, 获得基于pandpower的结果
        net.load.p_mw = pq_load[:net.load.shape[0]]
        net.load.q_mvar = pq_load[net.load.shape[0]:]
        try:
            pp.runopp(net, delta=1e-16, RETURN_RAW_DER=1)
        except:
            error_count += 1
            print(f"Error No.{error_count} happens in running the OPF.")
            continue
        new_QCQP = MPQCQP_Program(net=net, lb_loads=r_[p_load_init, q_load_init] * 0.8,
                                  ub_loads=r_[p_load_init, q_load_init] * 1.2)
        mu_ = new_QCQP.raw['lmbda']['mu']
        idx_act = find(mu_ > 1e-3).tolist()  # 修改成了 1e-3, 因为对互补性的阈值是1e-6，阈值=z*mu/(1+max(x)) => z*mu<5e-6;mu>1e-3能够保证z<1e-3
        # 将idx_act 分类储存起来，根据idx_act的长度和其中的内容，并对每种类型计数
        idx_act_tot = set(idx_act_tot) | set(idx_act)

        key = len(idx_act)
        if key not in all_data_lists_dict:
            all_data_lists_dict[key] = []

        if tuple(idx_act) in idx_act_dict.keys():
            idx_act_dict[tuple(idx_act)] = idx_act_dict.get(tuple(idx_act), 0) + 1
        else:
            idx_act_dict[tuple(idx_act)] = 1

        print(f"idx_act = {idx_act[:5]}")
        # 将idx_act 储存起来。
        # 如果idx_act与之前的所有idx_act都不同，那么就将这个idx_act加入到idx_act_list中；否则，continue
        idx_act = torch.tensor(idx_act).to(device)
        skip_current_iteration = False  # 设置标志
        for data_dict in all_data_lists_dict[key]:
            if torch.equal(data_dict['idx_act'], idx_act):  # 检查是否有重复；先只保留三个不等式约束被激活的情况
                skip_current_iteration = True  # 如果有重复，设置标志并跳出for循环
                break         # 如果有重复，那么就跳过这个idx_act

        if skip_current_iteration:  # 在for循环外检查标志
            continue  # 如果标志为True，跳过while循环的当前迭代

        add_data_per_type(all_data_lists_dict[key], idx_act, new_QCQP, pq_load)
    # 判断idx_act 是不是ineq_keep_del的子集
    if not set(idx_act_tot).issubset(set(ineq_keep_del)):
        #输出两个集合的差集 TODO: 这一块儿在进化学习的过程中也要补充
        print(f"idx_act_tot - ineq_keep_del = {set(idx_act_tot) - set(ineq_keep_del)}")

    # 输出idx_act_dict
    print(f"idx_act_dict = {idx_act_dict}")

    # 使用这个列表创建数据集实例dataset, 并将其保存到磁盘
    sorted_values = [all_data_lists_dict[key] for key in sorted(all_data_lists_dict.keys())]
    dataset = QCQPDataset(sorted_values)
    full_path = save_dataset_exl(Add_tot, dataset, idx_act_dict, unique=False, save=True)
    if len(dataset) == 0 or full_path is None:
        print(f"The dataset is empty. Exiting the program.{full_path}")
        sys.exit(0)  # Exit the program

    # # 加载数据集的时候，需要知道具体的文件名或者使用相同的格式来找到它
    # loaded_dataset = torch.load(full_path)
    #
    # batch_size = 16
    # for item in loaded_dataset:
    #     train_dataloader = DataLoader(item, batch_size=batch_size, shuffle=True)
    #
    #     # %%  4. 输出结果
    #     for batch_idx, batch_data in enumerate(train_dataloader):
    #         print(f"Batch {batch_idx + 1}:")
    #         # optimal_control_law_torch(batch_data)
    #
    #         cri_region = gen_cr_from_active_set_torch(batch_data)  # 生成当前的cr, 从GPU转移到CPU, 没使用keep_ineq_indices以减少计算量
    #     #
    #         for key, value in batch_data.items():
    #             # 选取每一批数据前5个示例
    #             example_values = value[:3]
    #
    #             # 打印键值对
    #             for example_idx, example_value in enumerate(example_values):
    #                 if key == 'idx_act':
    #                     print(f"{key} (Example {example_idx + 1}): {example_value}")
    #
    #         print("-" * 50)  # 分隔线

    #     for key, value in batch_data.items():
    #         # 由于我们是从DataLoader中获取数据，即使batch_size为1，返回的数据也会有一个额外的维度。
    #         # 使用squeeze()来去掉这个大小为1的维度。
    #         value = value.squeeze(0)
    #
    #         # 打印键值对
    #         print(f"{key}: {value}")
    #
    #     print("-" * 50)  # 分隔线
    #
    # for data in train_dataloader:
    #     if not opf_anly_sol.get_region(data['pq_load']):   # 批量的数据处理还没写
    #         print('pq_load is not inside the current critical region.')
    #
    #     cri_region = gen_cr_from_active_set_torch(data)  # 生成当前的critical region
    #     opf_anly_sol.add_region(cri_region)

    # 设置结束的时间，并输出总的消耗时间
    end_time = time.time()
    print(f"The total time is {end_time - start_time}")
    # print(f"The number of critical regions is {len(opf_anly_sol.critical_regions)}\nError_count = {error_count}")


def add_data_per_type(all_data_lists, idx_act, new_QCQP, pq_load):
    dh_dense = new_QCQP.dh.toarray()  # 将coo_matrix转换为dense numpy数组
    dg_dense = new_QCQP.dg.toarray()  # 将coo_matrix转换为dense numpy数组
    Lxx_dense = new_QCQP.Lxx.toarray()  # 将稀疏矩阵转换为dense numpy数组
    A_dense = new_QCQP.A.toarray()  # 将稀疏矩阵转换为dense numpy数组
    b_dense = new_QCQP.b  # 将稀疏矩阵转换为dense numpy数组
    A_t_dense = new_QCQP.A_t  # 将稀疏矩阵转换为dense numpy数组
    b_t_dense = new_QCQP.b_t  # 将稀疏矩阵转换为dense numpy数组
    Pd = new_QCQP.parameters_grid[0]
    idx_load = find(Pd != 0)
    # 储存所有的结构性数据. 将数据转换为PyTorch张量，并将其移动到GPU
    data_dict = {
        'pq_load': torch.tensor(pq_load).to(device),
        'idx_act': idx_act,
        'idx_load': torch.tensor(idx_load).to(device),
        'num_eq_conts': torch.tensor(new_QCQP.num_equality_constraints()).to(device),
        'num_ineq_conts': torch.tensor(new_QCQP.num_inequality_constraints()).to(device),
        'mu': torch.tensor(new_QCQP.raw['lmbda']['mu']).to(device),
        'lam': torch.tensor(new_QCQP.raw['lmbda']['lam']).to(device),
        'dh': torch.tensor(dh_dense).to(device),
        'dg': torch.tensor(dg_dense).to(device),
        'Lxx': torch.tensor(Lxx_dense).to(device),
        'Pd': torch.tensor(new_QCQP.parameters_grid[0]).to(device),
        'Qd': torch.tensor(new_QCQP.parameters_grid[1]).to(device),
        'F': torch.tensor(new_QCQP.F).to(device),
        'A': torch.tensor(A_dense).to(device),
        'b': torch.tensor(b_dense).to(device),
        'A_t': torch.tensor(A_t_dense).to(device),
        'b_t': torch.tensor(b_t_dense).to(device),
        'nb': torch.tensor(new_QCQP.nb).to(device),
        'x_0': torch.tensor(new_QCQP.x_0).to(device)
    }
    all_data_lists.append(data_dict)


if __name__ == "__main__":
    main()

